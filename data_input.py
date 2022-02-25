# -*- coding: utf-8 -*-
"""
Data Input
02/10/2022
Updates:
    - Convolutional Neural Network (CNN)
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
# All scikit-learn related modules
from sklearn.model_selection import train_test_split
import sklearn.metrics as skm
# All TensorFlow related modules
import tensorflow as tf
# import tensorflow_addons as tfa
from tensorflow.keras import layers, models
# import pathlib
from openpyxl import load_workbook
import datetime
import random
import time
import os
import sys

# Will need to have Input Data as 3-dimensional ILI layers
# - var1: number of ILI data files
# - var2: dimension in the longitudinal (along pipe axis)
# - var3: dimension in the lateral (along pipe circumference)
# To simplify the preprocessing, will use square resolution, meaning that the
# unit of length in var2 should be equal to that of var3. This way the length
# is not warped in any direction (more resolution in one direction). However,
# this does not mean that values var2 and var3 need to be equal. The resultant
# image should be like a rectangle.

# =============================================================================
# PARAMETERS
# =============================================================================

time_start = time.time()

# Input Data dimensions
axial_len = 200
circ_len = 200
# Multicategorical classifier = TRUE, Binary classifier = FALSE
multicat_class = True

# Parameter Options
parameter_1 = 'NA'
parameter_2 = 0.2
# p3 = ['relu', 'sigmoid','tanh']
# ['relu','tanh']
p3 = ['relu','tanh']
p4 = ['relu','tanh']
p5 = ['relu','tanh']
p6 = ['relu','tanh']
p7 = ['adam','SGD','RMSprop']
parameter_8 = 'accuracy'
p9 = [tf.keras.losses.SparseCategoricalCrossentropy(from_logits=True),
               tf.keras.losses.BinaryCrossentropy(from_logits=True)]
parameter_10 = ''

run_inf = True
while run_inf:
    # parameter_3 = p3[random.randint(0,len(p3)-1)]
    # parameter_4 = p4[random.randint(0,len(p4)-1)]
    # parameter_5 = p5[random.randint(0,len(p5)-1)]
    # parameter_6 = p6[random.randint(0,len(p6)-1)]
    # parameter_7 = p7[random.randint(0,len(p7)-1)]
    parameter_3 = p3[0]
    parameter_4 = p4[0]
    parameter_5 = p5[0]
    parameter_6 = p6[0]
    parameter_7 = p7[0]
    
    # parameter_9 = p9[random.randint(0,len(p9)-1)]
    parameter_9 = p9[0]
    
    # Current Time
    time_total = time.time() - time_start
    time_m, time_s = divmod(time_total, 60)
    time_h, time_m = divmod(time_m, 60)
    print('========== START ==========')
    print('Total execution time of %.0f:%.0f:%.0f' % (time_h, time_m, time_s))    
    time_iteration = time.time()
    
    # # Parameter Organization
    # parameter_1 = 6         # train_test_split: random_state
    # parameter_2 = 0.2       # train_test_split: test_size
    # parameter_3 = 'relu'    # activation for first layer
    # parameter_4 = 'relu'    # activation for second layer
    # parameter_5 = 'relu'    # activation for third layer
    # parameter_6 = 'relu'    # activation for dense layer
    # parameter_7 = 'adam'    # optimizer for model
    # parameter_8 = 'accuracy' # metrics for model
    # parameter_9 = ''
    # parameter_10 = ''
    
    # =============================================================================
    # IMAGE AND LABEL DATA INPUT
    # =============================================================================
    
    # Load Training Data
    training_data_path = 'training_data'
    
    image_data_path = []
    label_data_path = []
    
    for subdir, dirs, files in os.walk(training_data_path):
        for file in files:
            if 'radius' in file:
                image_data_path.append(os.path.join(subdir, file))
            if 'SCF' in file:
                label_data_path.append(os.path.join(subdir, file))
                
    # Load ALL of the data .npy files
    if len(image_data_path) != len(label_data_path):
        sys.exit()
    
    for i in range(len(image_data_path)):
        if i == 0: 
            image_data = np.load(image_data_path[i])
            label_data = np.load(label_data_path[i])
            continue
        image_data = np.concatenate((image_data, np.load(image_data_path[i])), axis=0)
        label_data = np.concatenate((label_data, np.load(label_data_path[i])), axis=0)
        
    # image_data = np.load(image_data_path)
    # label_data = np.load(label_data_path)
    label_data_cat = np.zeros(len(label_data),dtype=int)
    
    # =============================================================================
    # CLASSES
    # =============================================================================
    
    if multicat_class == True:
        # Multicategorical test data
        class_names = ['Low','Low-Mid','Mid','Mid-High','High']
        class_count = 5
        c1 = 1.5
        c2 = 2
        c3 = 3
        c4 = 4
        for i, val in enumerate(label_data):
            if val < c1: 
                label_data_cat[i] = 0
            elif val >= c1 and val < c2: 
                label_data_cat[i] = 1
            elif val >= c2 and val < c3:
                label_data_cat[i] = 2
            elif val >= c3 and val < c4:
                label_data_cat[i] = 3
            elif val >= c4:
                label_data_cat[i] = 4
    else:
        # Binary categorical test data
        class_names = ['Low','High']
        class_count = 2
        # Choosing a c1 value of 2.6 gives just about an equal number of data samples
        # above and below the threshold. This is true as of 2/10/2022
        c1 = 3
        c2 = 'NA'
        c3 = 'NA'
        c4 = 'NA'
        for i, val in enumerate(label_data):
            if val < c1: 
                label_data_cat[i] = 0
            elif val >= c1: 
                label_data_cat[i] = 1
    
        
    # Split the data into Training and Test sets
    random_state = parameter_1
    test_size = parameter_2
    # Using random state 6 for now because it ensures that out of the 26 data points, the test_labels will contain
    # at least one of each of the 5 class types
    train_images, test_images = train_test_split(image_data, test_size=test_size)#, random_state=random_state)
    train_labels, test_labels = train_test_split(label_data_cat, test_size=test_size)#, random_state=random_state)
    
    # Test 9 random images
    # Generate a random set of test images and labels
    rand_list_test = random.sample(range(train_images.shape[0]), 9)
    plt.figure(figsize=(10,10))
    for i, j in enumerate(rand_list_test):
        plt.subplot(3,3,i+1)
        plt.xticks([])
        plt.yticks([])
        plt.grid(False)
        plt.imshow(train_images[j], cmap=plt.cm.RdYlGn)#, vmin=-0.02, vmax=0.02)#cmap=plt.cm.binary, vmin=-0.01, vmax=0.01)
        plt.xlabel(class_names[train_labels[j]])
        plt.colorbar()
    plt.show()
    
    # Create the convolutional base
    model = models.Sequential()
    model.add(layers.Conv2D(32, (3,3), activation=parameter_3, input_shape=(axial_len,circ_len,1)))
    model.add(layers.MaxPooling2D((2,2)))
    model.add(layers.Conv2D(64, (3,3), activation=parameter_4))
    model.add(layers.MaxPooling2D((2,2)))
    model.add(layers.Conv2D(128, (3,3), activation=parameter_5))
    
    # Add dense layers on top
    model.add(layers.Flatten())
    model.add(layers.Dense(200, activation=parameter_6))
    model.add(layers.Dense(class_count))
    
    # Compile and train model
    model.compile(optimizer=parameter_7,
                  # loss=tf.keras.losses.SparseCategoricalCrossentropy(from_logits=True),
                  loss=parameter_9,
                  metrics=[parameter_8])
    
    training_time_start = time.time()
    history = model.fit(train_images, train_labels, epochs=10,
                        validation_data=(test_images, test_labels))
    training_time_stop = time.time()
    
    # Evaluate the model
    plt.figure(figsize=(8,6))
    plt.plot(history.history['accuracy'], label='Accuracy')
    plt.plot(history.history['val_accuracy'], label='Val. Accuracy')
    plt.xlabel('Epoch')
    plt.ylabel('Accuracy')
    plt.ylim([0, 1])
    plt.legend(loc='lower right')
    
    test_loss, test_acc = model.evaluate(test_images, test_labels, verbose=2)
    
    # Make predictions
    probability_model = tf.keras.Sequential([model,
                                              tf.keras.layers.Softmax()])
    predictions = probability_model.predict(test_images)
    
    # Graph for set of 10 class predictions
    def plot_image(i, predictions_array, true_label, img):
        true_label, img = true_label[i], img[i]
        plt.grid(False)
        plt.xticks([])
        plt.yticks([])
        
        plt.imshow(img, cmap=plt.cm.RdYlGn)
        
        predicted_label = np.argmax(predictions_array)
        
        if predicted_label == true_label:
            color = 'blue'
        else:
            color = 'red'
        
        plt.xlabel("Dent#{}: {} {:2.0f}% ({})".format(i,
                                             class_names[predicted_label],
                                              100*np.max(predictions_array),
                                              class_names[true_label]),
                    color=color)
        
    def plot_value_array(i, predictions_array, true_label):
        true_label = true_label[i]
        plt.grid(False)
        plt.xticks(range(class_count))
        plt.yticks([])
        thisplot = plt.bar(range(class_count), predictions_array, color="#777777")
        plt.ylim([0,1])
        
        predicted_label = np.argmax(predictions_array)
        
        thisplot[predicted_label].set_color('red')
        thisplot[true_label].set_color('blue')
        
    # Plot several images
    num_rows = 3
    num_cols = 3
    num_images = num_rows * num_cols
    plt.figure(figsize=(2*2*num_cols, 2*num_rows))
    # Generate a random set of test images and labels
    rand_list = random.sample(range(predictions.shape[0]), num_images)
    
    for i, j in enumerate(rand_list):
        plt.subplot(num_rows, 2*num_cols, 2*i+1)
        plot_image(j, predictions[j], test_labels, test_images)
        plt.subplot(num_rows, 2*num_cols, 2*i+2)
        plot_value_array(j, predictions[j], test_labels)
    plt.tight_layout()
    plt.show()
    
    # Use the trained model to make predictions
    j = 3
    def single_test(j):
        img = test_images[j]
        img = (np.expand_dims(img,0))
        predictions_single = probability_model.predict(img)
        plt.figure(figsize=(6,3))
        plt.subplot(1,2,1)
        plot_image(j, predictions_single, test_labels, test_images)
        plt.subplot(1,2,2)
        plot_value_array(j, predictions[j], test_labels)
        plt.tight_layout()
        plt.show()
        
    # =============================================================================
    # METRICS
    # =============================================================================
    
    y_test = test_labels.copy()
    y_pred = [np.argmax(i) for i in predictions]
    y_pred_prob = predictions.copy()
    
    # Classification Report
    # metrics_class_report = skm.classification_report(y_test, y_pred)
    metrics_class_report = skm.classification_report(y_test, y_pred, target_names = class_names, output_dict=True)
    
    metrics_class_report_print = skm.classification_report(y_test, y_pred, target_names = class_names)
    print(metrics_class_report_print)
    
    # print(metrics_class_report)
    # ROC AUC Score
    if multicat_class == True:
        metrics_roc_auc = skm.roc_auc_score(y_test, y_pred_prob, average="weighted", multi_class="ovr")
    else:
        y_pred_prob_binary = [np.max(i) for i in y_pred_prob]
        metrics_roc_auc = skm.roc_auc_score(y_test, y_pred_prob_binary)
    # Cohen's Kappa Score
    metrics_kappa = skm.cohen_kappa_score(y_test, y_pred)
    # Matthew's correlation coefficient (MCC)
    metrics_mcc = skm.matthews_corrcoef(y_test, y_pred)
    # Log Loss
    metrics_log_loss = skm.log_loss(y_test, y_pred_prob)
    
    # metric = tfa.metrics.F1Score(num_classes=5)
    # metric.update_state(y_test, y_pred_prob)
    
    df_data = [metrics_roc_auc, metrics_kappa, metrics_mcc, metrics_log_loss]
    df_labels = ['ROC AUC', 'Kappa', 'MCC', 'Log Loss']
    df = pd.DataFrame(data=df_data, index=df_labels, columns=['Value'])
    print(df)
    
    # Save the SCF value to an Excel sheet
    # Save the Metrics to the metrics.xlsx Workbook
    metrics_path = 'metrics/'
    metrics_name = 'metrics.xlsx'
    metrics_file = metrics_path + metrics_name
    wbM = load_workbook(metrics_file)#, data_only=True)
    wbM_sn = wbM.sheetnames
    wbMs = wbM[wbM_sn[0]]
    wbMs_row = wbMs.max_row
    
    metrics_labels = wbMs[2]
    metrics_labels = [cell.value for cell in metrics_labels]
    
    # Metrics Values
    time_stamp = str(datetime.datetime.now())
    run_number = wbMs_row - 1
    training_size = train_labels.shape[0]
    training_size_0 = sum(1 for i in train_labels if i == 0)
    training_size_1 = sum(1 for i in train_labels if i == 1)
    training_size_2 = sum(1 for i in train_labels if i == 2)
    training_size_3 = sum(1 for i in train_labels if i == 3)
    training_size_4 = sum(1 for i in train_labels if i == 4)
    training_accuracy = history.history['accuracy'][-1]
    training_time = training_time_stop - training_time_start
    if multicat_class == True:
        precision_0 = metrics_class_report[class_names[0]]['precision']
        precision_1 = metrics_class_report[class_names[1]]['precision']
        precision_2 = metrics_class_report[class_names[2]]['precision']
        precision_3 = metrics_class_report[class_names[3]]['precision']
        precision_4 = metrics_class_report[class_names[4]]['precision']
        recall_0 = metrics_class_report[class_names[0]]['recall']
        recall_1 = metrics_class_report[class_names[1]]['recall']
        recall_2 = metrics_class_report[class_names[2]]['recall']
        recall_3 = metrics_class_report[class_names[3]]['recall']
        recall_4 = metrics_class_report[class_names[4]]['recall']
    else:
        precision_0 = metrics_class_report[class_names[0]]['precision']
        precision_1 = metrics_class_report[class_names[1]]['precision']
        precision_2 = 'NA'
        precision_3 = 'NA'
        precision_4 = 'NA'
        recall_0 = metrics_class_report[class_names[0]]['recall']
        recall_1 = metrics_class_report[class_names[1]]['recall']
        recall_2 = 'NA'
        recall_3 = 'NA'
        recall_4 = 'NA'
    f1 = metrics_class_report['weighted avg']['f1-score']
    roc_auc = metrics_roc_auc
    kappa = metrics_kappa
    mcc = metrics_mcc
    log_loss = metrics_log_loss
    accuracy = metrics_class_report['accuracy']
    total_time = time.time() - time_iteration
    
    # Metrics Values in list organization
    metrics_data = [time_stamp,
                    run_number,
                    class_count,
                    c1,
                    c2,
                    c3,
                    c4,
                    parameter_1,
                    parameter_2,
                    parameter_3,
                    parameter_4,
                    parameter_5,
                    parameter_6,
                    parameter_7,
                    parameter_8,
                    str(parameter_9),
                    parameter_10,
                    training_size,
                    training_size_0,
                    training_size_1,
                    training_size_2,
                    training_size_3,
                    training_size_4,
                    training_accuracy,
                    training_time,
                    precision_0,
                    precision_1,
                    precision_2,
                    precision_3,
                    precision_4,
                    recall_0,
                    recall_1,
                    recall_2,
                    recall_3,
                    recall_4,
                    f1,
                    roc_auc,
                    kappa,
                    mcc,
                    log_loss,
                    accuracy,
                    total_time]
    
    # Write the metric values to the Excel workbook
    for i, value in enumerate(metrics_data):
        wbMs.cell(row=wbMs_row+1, column=i+1, value=value)
    wbM.save(metrics_file)
    # Close the Results Workbook
    wbM.close()
    
    # Total time
    time_total = time.time() - time_start
    time_m, time_s = divmod(time_total, 60)
    time_h, time_m = divmod(time_m, 60)
    print('Total execution time of %.0f:%.0f:%.0f' % (time_h, time_m, time_s))
    print('=========== END ===========')