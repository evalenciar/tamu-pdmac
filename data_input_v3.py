# -*- coding: utf-8 -*-
"""
Data Input
02/10/2022
Updates:
    - Convolutional Neural Network (CNN)
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
# All scikit-learn related modules
from sklearn.model_selection import train_test_split
import sklearn.metrics as skm
# All TensorFlow related modules
import tensorflow as tf
# import tensorflow_addons as tfa
from tensorflow.keras import layers, models
# import pathlib
from openpyxl import load_workbook
import datetime
import random
import time
import os
import sys

# Will need to have Input Data as 3-dimensional ILI layers
# - var1: number of ILI data files
# - var2: dimension in the longitudinal (along pipe axis)
# - var3: dimension in the lateral (along pipe circumference)
# To simplify the preprocessing, will use square resolution, meaning that the
# unit of length in var2 should be equal to that of var3. This way the length
# is not warped in any direction (more resolution in one direction). However,
# this does not mean that values var2 and var3 need to be equal. The resultant
# image should be like a rectangle.

# =============================================================================
# PARAMETERS
# =============================================================================

time_start = time.time()

# Input Data dimensions
axial_len = 200
circ_len = 200
shape = [axial_len, circ_len]

# Multicategorical classifier = TRUE, Binary classifier = FALSE
class_binary_bool = True
class_threshold = [3]
class_count = len(class_threshold) + 1
# class_names = ['Low','Low-Mid','Mid','Mid-High','High']
class_names = ['Low','High']

# Parameter Options
parameter_1 = 'NA'
parameter_2 = 0.2
# p3 = ['relu', 'sigmoid','tanh']
# ['relu','tanh']
p3 = ['relu','tanh']
p4 = ['relu','tanh']
p5 = ['relu','tanh']
p6 = ['relu','tanh']
p7 = ['adam','SGD','RMSprop']
parameter_8 = 'accuracy'
p9 = [tf.keras.losses.SparseCategoricalCrossentropy(from_logits=True),
               tf.keras.losses.BinaryCrossentropy(from_logits=True)]
parameter_10 = ''

# =============================================================================
# FUNCTIONS
# =============================================================================

def load_training_data(training_data_path):
    # Load Training Data
    image_data_path = []
    label_data_path = []

    for subdir, dirs, files in os.walk(training_data_path):
        for file in files:
            if 'radius' in file:
                image_data_path.append(os.path.join(subdir, file))
            if 'SCF' in file:
                label_data_path.append(os.path.join(subdir, file))
                
    # Load ALL of the data .npy files
    if len(image_data_path) != len(label_data_path):
        sys.exit()

    for i in range(len(image_data_path)):
        if i == 0: 
            image_data = np.load(image_data_path[i])
            label_data = np.load(label_data_path[i])
            continue
        
        image_data = np.concatenate((image_data, np.load(image_data_path[i])), axis=0)
        label_data = np.concatenate((label_data, np.load(label_data_path[i])), axis=0)
        
    return image_data, label_data

def class_definition(label_data, class_binary_bool, class_threshold):
    # Initiate the categorical label array
    label_data_cat = np.zeros(len(label_data),dtype=int)
    
    if class_binary_bool == True:
        # Binary categorical test data
        
        # Choosing a c1 value of 2.6 gives just about an equal number of data samples
        # above and below the threshold. This is true as of 2/10/2022
        ct = class_threshold[0]
        for i, val in enumerate(label_data):
            if val < ct: 
                label_data_cat[i] = 0
            elif val >= ct: 
                label_data_cat[i] = 1
    else:
        # Multicategorical test data
        ct = class_threshold
                
        for i, val in enumerate(label_data):
            for j in range(0,len(ct)):
                if val < ct[j]:
                    label_data_cat[i] = j
                    break
                elif val > ct[-1]:
                    label_data_cat[i] = ct[-1]

    return label_data_cat

def regressor_model(shape):
    model = models.Sequential()
    model.add(layers.Conv2D(32, (3,3), activation='relu', input_shape=(shape[0], shape[1], 1)))
    model.add(layers.MaxPooling2D((2,2)))
    model.add(layers.Conv2D(64, (3,3), activation='relu'))
    model.add(layers.MaxPooling2D((2,2)))
    model.add(layers.Conv2D(128, (3,3), activation='relu'))
    model.add(layers.MaxPooling2D((2,2)))
    model.add(layers.Conv2D(256, (3,3), activation='relu'))
    model.add(layers.MaxPooling2D((2,2)))
    
    # Add dense layers on top
    model.add(layers.Flatten())
    model.add(layers.Dense(100, activation='relu'))
    model.add(layers.Dense(10, activation='relu'))
    model.add(layers.Dense(1))
    
    # Compile and train model
    model.compile(optimizer='RMSprop', loss='mse', metrics=['mae'])
    
    return model

def plot_image(i, predictions_array, true_label, img):
    # Graph for set of 10 class predictions
    true_label, img = true_label[i], img[i]
    plt.grid(False)
    plt.xticks([])
    plt.yticks([])
    plt.imshow(img, cmap=plt.cm.RdYlGn)
    predicted_label = np.argmax(predictions_array)
    if predicted_label == true_label:
        color = 'blue'
    else:
        color = 'red'
    
    plt.xlabel("Dent#{}: {} {:2.0f}% ({})".format(i, 
                                                  class_names[predicted_label],
                                                  100*np.max(predictions_array),
                                                  class_names[true_label]),
               color=color)

def plot_value_array(i, predictions_array, true_label):
    true_label = true_label[i]
    plt.grid(False)
    plt.xticks(range(class_count))
    plt.yticks([])
    thisplot = plt.bar(range(class_count), predictions_array, color="#777777")
    plt.ylim([0,1])
    predicted_label = np.argmax(predictions_array)
    thisplot[predicted_label].set_color('red')
    thisplot[true_label].set_color('blue')

def single_prediction(j):
    img = images_test[j]
    img = (np.expand_dims(img,0))
    predictions_single = probability_model.predict(img)
    plt.figure(figsize=(6,3))
    plt.subplot(1,2,1)
    plot_image(j, predictions_single, labels_test, images_test)
    plt.subplot(1,2,2)
    plot_value_array(j, predictions[j], labels_test)
    plt.tight_layout()
    plt.show()

def metrics(class_binary_bool):
    # Classification Report
    class_report = skm.classification_report(y_test, y_pred, target_names = class_names, output_dict=True)
    print(skm.classification_report(y_test, y_pred, target_names = class_names))
    # ROC AUC Score
    if class_binary_bool == True:
        y_pred_prob_binary = [np.max(i) for i in y_pred_prob]
        roc_auc = skm.roc_auc_score(y_test, y_pred_prob_binary)
    else:
        roc_auc = skm.roc_auc_score(y_test, y_pred_prob, average="weighted", multi_class="ovr")
    
    # Cohen's Kappa Score
    kappa = skm.cohen_kappa_score(y_test, y_pred)
    # Matthew's correlation coefficient (MCC)
    mcc = skm.matthews_corrcoef(y_test, y_pred)
    # Log Loss
    log_loss = skm.log_loss(y_test, y_pred_prob)
    
    return class_report, roc_auc, kappa, mcc, log_loss

def classifier_metrics_array(class_binary_bool):
    # Metrics Values
    training_size_class = []
    for i in range(0, class_count):
        training_size_class.append(sum(1 for j in labels_train if j == i))

    if class_binary_bool == True:
        precision_0 = metrics_class_report[class_names[0]]['precision']
        precision_1 = metrics_class_report[class_names[1]]['precision']
        precision_2 = 'NA'
        precision_3 = 'NA'
        precision_4 = 'NA'
        recall_0 = metrics_class_report[class_names[0]]['recall']
        recall_1 = metrics_class_report[class_names[1]]['recall']
        recall_2 = 'NA'
        recall_3 = 'NA'
        recall_4 = 'NA'
    else:        
        precision_0 = metrics_class_report[class_names[0]]['precision']
        precision_1 = metrics_class_report[class_names[1]]['precision']
        precision_2 = metrics_class_report[class_names[2]]['precision']
        precision_3 = metrics_class_report[class_names[3]]['precision']
        precision_4 = metrics_class_report[class_names[4]]['precision']
        recall_0 = metrics_class_report[class_names[0]]['recall']
        recall_1 = metrics_class_report[class_names[1]]['recall']
        recall_2 = metrics_class_report[class_names[2]]['recall']
        recall_3 = metrics_class_report[class_names[3]]['recall']
        recall_4 = metrics_class_report[class_names[4]]['recall']

    # Metrics Values in list organization
    metrics_data = [str(datetime.datetime.now()),   # Time stamp
                    wbMs_row - 1,                   # Run number
                    class_count,
                    str(class_threshold),           # Array of class thresholds
                    'NA',
                    'NA',
                    'NA',
                    parameter_1,                    # train_test_split: random_state
                    parameter_2,                    # train_test_split: test_size
                    parameter_3,                    # activation for first layer
                    parameter_4,                    # activation for second layer
                    parameter_5,                    # activation for third layer
                    parameter_6,                    # activation for dense layer
                    parameter_7,                    # optimizer for model
                    parameter_8,                    # metrics for model
                    str(parameter_9),
                    parameter_10,
                    labels_train.shape[0],          # Training size
                    str(training_size_class),       # Array of number of data points per class
                    'NA',
                    'NA',
                    'NA',
                    'NA',
                    history.history['accuracy'][-1],            # Training accuracy
                    training_time_stop - training_time_start,   # Training time
                    precision_0,
                    precision_1,
                    precision_2,
                    precision_3,
                    precision_4,
                    recall_0,
                    recall_1,
                    recall_2,
                    recall_3,
                    recall_4,
                    metrics_class_report['weighted avg']['f1-score'], # F1 score
                    metrics_roc_auc,
                    metrics_kappa,
                    metrics_mcc,
                    metrics_log_loss,
                    metrics_class_report['accuracy'],   # Accuracy
                    time.time() - time_iteration]       # Total time
    
    return metrics_data

# =============================================================================
# SCRIPT ITERATIONS
# =============================================================================

# run_inf = True
# while run_inf:
# parameter_3 = p3[random.randint(0,len(p3)-1)]
# parameter_4 = p4[random.randint(0,len(p4)-1)]
# parameter_5 = p5[random.randint(0,len(p5)-1)]
# parameter_6 = p6[random.randint(0,len(p6)-1)]
# parameter_7 = p7[random.randint(0,len(p7)-1)]
parameter_3 = p3[0]
parameter_4 = p4[0]
parameter_5 = p5[0]
parameter_6 = p6[0]
parameter_7 = p7[2]

# parameter_9 = p9[random.randint(0,len(p9)-1)]
parameter_9 = p9[0]

# Current Time
time_total = time.time() - time_start
time_m, time_s = divmod(time_total, 60)
time_h, time_m = divmod(time_m, 60)
print('========== START ==========')
print('Total execution time of %.0f:%.0f:%.0f' % (time_h, time_m, time_s))    
time_iteration = time.time()

image_data, label_data = load_training_data('training_data')

label_cat_data = class_definition(label_data, True, class_threshold)

# Split the data into Training and Test sets
images_train, images_test, labels_train, labels_test = train_test_split(image_data, label_data, test_size=0.2)
# images_train, images_test = train_test_split(image_data, test_size=test_size, random_state=random_state)
# labels_train, labels_test = train_test_split(label_data, test_size=test_size, random_state=random_state)


# Test 9 random images
# Generate a random set of test images and labels
rand_list_test = random.sample(range(images_train.shape[0]), 9)
plt.figure(figsize=(10,10))
for i, j in enumerate(rand_list_test):
    plt.subplot(3,3,i+1)
    plt.xticks([])
    plt.yticks([])
    plt.grid(False)
    plt.imshow(images_train[j], cmap=plt.cm.RdYlGn)
    plt.xlabel(labels_train[j])
    plt.colorbar()
plt.show()

model = regressor_model(shape)  # Create the regression ML model
model.summary()                 # Print the current model information

training_time_start = time.time()
history = model.fit(images_train, labels_train, epochs=50,
                    validation_data=(images_test, labels_test))
training_time_stop = time.time()

# Evaluate the model: Metrics
plt.figure(figsize=(8,6))
plt.subplot(2,1,1)
plt.plot(history.history['mae'], label='MAE')
plt.plot(history.history['val_mae'], label='Val. MAE')
plt.xlabel('Epoch')
plt.ylabel('MAE')
plt.ylim([0, 1])
plt.legend(loc='lower right')
# Evaluate the model: Losses
plt.subplot(2,1,2)
plt.plot(history.history['loss'], label='Loss')
plt.plot(history.history['val_loss'], label='Val. Loss')
plt.xlabel('Epoch')
plt.ylabel('Loss')
plt.ylim([0, 1])
plt.legend(loc='lower right')

test_loss, test_acc = model.evaluate(images_test, labels_test, verbose=2)

# =============================================================================
# VALIDATION
# =============================================================================

# Make predictions
probability_model = tf.keras.Sequential([model,
                                          tf.keras.layers.Softmax()])
predictions = probability_model.predict(images_test)

# Plot several images
num_rows = 3
num_cols = 3
num_images = num_rows * num_cols
plt.figure(figsize=(2*2*num_cols, 2*num_rows))
# Generate a random set of test images and labels
rand_list = random.sample(range(predictions.shape[0]), num_images)

for i, j in enumerate(rand_list):
    plt.subplot(num_rows, 2*num_cols, 2*i+1)
    plot_image(j, predictions[j], labels_test, images_test)
    plt.subplot(num_rows, 2*num_cols, 2*i+2)
    plot_value_array(j, predictions[j], labels_test)
plt.tight_layout()
plt.show()
    
# =============================================================================
# METRICS
# =============================================================================

y_test = labels_test.copy()
y_pred = [np.argmax(i) for i in predictions]
y_pred_prob = predictions.copy()

metrics_class_report, metrics_roc_auc, metrics_kappa, metrics_mcc, metrics_log_loss = metrics(class_binary_bool)
# Print the results to view
df_data = [metrics_roc_auc, metrics_kappa, metrics_mcc, metrics_log_loss]
df_labels = ['ROC AUC', 'Kappa', 'MCC', 'Log Loss']
df = pd.DataFrame(data=df_data, index=df_labels, columns=['Value'])
print(df)

# Save the SCF value to an Excel sheet
# Save the Metrics to the metrics.xlsx Workbook
metrics_path = 'metrics/'
metrics_name = 'metrics.xlsx'
metrics_file = metrics_path + metrics_name
wbM = load_workbook(metrics_file)
wbM_sn = wbM.sheetnames
wbMs = wbM[wbM_sn[0]]
wbMs_row = wbMs.max_row

metrics_labels = wbMs[2]
metrics_labels = [cell.value for cell in metrics_labels]

metrics_data = classifier_metrics_array(class_binary_bool)

# Write the metric values to the Excel workbook
for i, value in enumerate(metrics_data):
    wbMs.cell(row=wbMs_row+1, column=i+1, value=value)
wbM.save(metrics_file)
wbM.close()     # Close the Results Workbook

# Total time
time_total = time.time() - time_start
time_m, time_s = divmod(time_total, 60)
time_h, time_m = divmod(time_m, 60)
print('Total execution time of %.0f:%.0f:%.0f' % (time_h, time_m, time_s))
print('=========== END ===========')