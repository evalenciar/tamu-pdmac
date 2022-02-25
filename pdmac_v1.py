# -*- coding: utf-8 -*-
"""
PDMAC Tool
Version 0.8.2
Created on 08/09/2021
Updated by Emmanuel Valencia, 01/05/2022

This script is proprietary to ADV Integrity, Inc. It processes In-Line
Inspection (ILI) data to prepare for FEA analysis.
"""

# Import Excel and Word editing modules
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# from docx import Document
# from docx.shared import Inches
# from mailmerge import MailMerge
# Import math and array related modules
import numpy as np
import pandas as pd
import math
# Import data smoothing modules
from scipy.interpolate import splrep, splev
from scipy.signal import savgol_filter
import matplotlib.pyplot as plt
# Import system modules
import time
import os
# import shutil, os
# import readline
# import sys

# =============================================================================
# PDMAC PARAMETERS
# =============================================================================

# Turn off interactive plotting
plt.ioff()

# Data smoothing
# data_smoothing = False
degree = 4
refinement = 1
# Time limit of 5 minutes for searching for the .sta file
time_limit = 60*5
circ_grad_tol = 0.05

remove_calipers     = False
inp_file_bool       = True
inp_file_ext        = 's'
create_graphs       = True
submit_to_abaqus    = True
output_report       = True

if inp_file_ext == 'r':
    data_smoothing = False
    isolate_dent = False
elif inp_file_ext == 's':
    data_smoothing = True
    isolate_dent = False
elif inp_file_ext == 'i':
    data_smoothing = True
    isolate_dent = True

# ----- Iterate through all of the ILI data files -----

data_path = 'raw_data'
overall_time = time.time()

first_file = True

# Load the SCF Values DataFrame
scf_df = pd.read_excel('scf_values.xlsx')

for subdir, dirs, files in os.walk(data_path):
    # Ignore folder
    if 'Ignore' in subdir:
        continue
    # Currently looping through X folder
    print('OT-%03d | ======= PDMAC START ========' % (time.time() - overall_time))
    print('OT-%03d | Began processing subdir: %s' % (time.time() - overall_time,subdir))
    
    # Find the corresponding Dent Registry (or Tally Sheet)
    # For now, assume that the Dent Registry is ALWAYS an .xlsm file
    for file in files:
        if not file.find('.xlsm') == -1:
            print('OT-%03d | Found Batch Tally Sheet: %s' % (time.time() - overall_time,file))
            reg_path = os.path.join(subdir,file)
            reg = pd.read_excel(reg_path, sheet_name=0, header=8)
    
    for file in files:
        if not file.find('.xlsm') == -1:
            continue
        print('OT-%03d | ========== START ===========' % (time.time() - overall_time))
        print('OT-%03d | Began processing file: %s' % (time.time() - overall_time,file))
        # print(os.path.join(subdir,file))
        data_file = os.path.join(subdir,file)
            
# =============================================================================
# DATA PROCESSING STEP 1: LOAD DATA AND DENT REGISTRY
# =============================================================================
        # Begin tracking time
        start_time = time.time()
        
        # File name and relative location
        # data_file = data_path + '/' + data_name.name
        
        # Check if the file is a .csv file
        if not file.find('.csv') == -1:
            
            # rawdata = pd.read_csv(data_file, header=None)   #open current csv file into rawdata
            # rawdata = np.array(rawdata)                     #[calipers, axial position]
            # z_start = rawdata[0,1]                          #start z-value
        
            # rotated_data = np.zeros((rawdata.shape[0] - 1,rawdata.shape[1]))
            # rotated_data[0,:] = rawdata[0,1:]   # Use first row as z-values
            # rawdata = rawdata[1:,1:]            # Remove first column and row
            # min_loc = np.unravel_index(np.argmin(rawdata), rawdata.shape)
            
            # rotated_half = rawdata.shape[0]/2
            # rotated_start = int(min_loc[0] - rotated_half + 1)
            
            # for line, rad in enumerate(rawdata[:,0]): #loop to rotate data
            #     if line + rotated_start < rotated_data.shape[0]-1:
            #         rotated_data[line + 1,:]=rawdata[line + rotated_start,:]
            #     if line + rotated_start >= rotated_data.shape[0]-1:
            #         rotated_data[line + 1,:]=rawdata[line + rotated_start - rotated_data.shape[0] + 1,:]
        
            rawdata = pd.read_csv(data_file, header=None)   #open current csv file into rawdata
            rawdata = np.array(rawdata)                     #[calipers, axial position]
            z_start = rawdata[0,1]   
        
            data         = rawdata[:, 1:rawdata.shape[1]:1] #remove first column
            rotated_data = np.zeros(data.shape)             #size rotated array after first column is removed
            data         = data[1:data.shape[0]:1,:]        #remove first row
            min_loc      = np.where(data == np.min(data))   #find location of minimum value
        
            half    = data.shape[0]/2                 #find half of the data points
            start   = int(min_loc[0][0] - half + 1)   #account for arrays starting at zero
            rotated_data[0,:] = rawdata[0][1:]        #take first column as axial distances
        
            for line, rad in enumerate(data[:,0]): #loop to rotate data
                if line+start < rotated_data.shape[0]-1:
                    rotated_data[line+1,:]=data[line+start,:]
                if line+start >= rotated_data.shape[0]-1:
                    rotated_data[line+1,:]=data[line+start-rotated_data.shape[0]+1,:]
            
            # ----- Continue using Pandas DataFrame -----
            theta_i     = np.arange(rotated_data.shape[0] - 1)*360/(rotated_data.shape[0]-1)
            # theta_i     = data[1:,0].astype(float)*360/(len(data[1:,0]))
            radius_i    = rotated_data[1:,:]
            # radius_i    = data[1:,1:].astype(float)
            # z_i         = data[0,1:].astype(float)
            z_i         = np.array([(s - z_start)*12 for s in rotated_data[0,:].astype(float)])
            rows_i      = theta_i.shape[0]
            cols_i      = z_i.shape[0]
            
            
        else:
        
            # ----- Begin using OpenPyXL to do the following -----
            # - Locate the first highlighted cell in col A indicating deg 0
            # - Rearrange the rows so that the firs↓t row is deg 0
            # - Convert the Column A caliper number data into theta values, from 0 to 360
            # - Delete repeating columns based on the z value
            # - Convert the z values from feet to inches
            wb = load_workbook(data_file, data_only = True)
            wb_sn = wb.sheetnames
            wbs = wb[wb_sn[0]]
            
            # Different vendors use different formatting:
            # - Some vendors highlight certain caliper cells to indicate the 0 deg caliper
            # Check if a value exists on cell B1
            if (not bool(wbs['B1'].value)) or (bool(wbs['B1'].value) and isinstance(wbs['B1'].value, str)):
            
                # Determine the caliper sensor spacing as recorded in the Excel file
                caliper_spacing = wbs['A1'].value
                # Assuming that the value has 3 decimal places, use 5
                caliper_spacing = float(caliper_spacing[-5:wbs['A1'].value.find('"')])
                
                # Unmerge the cells in the first row before deleting
                for cell_group in wbs.merged_cells.ranges:
                    wbs.unmerge_cells(str(cell_group))
                    
                # Delete the first row since no longer necessary
                wbs.delete_rows(1,1)
                
                # Find the yellow highlighted cell in column A, this is 0deg
                highlight_color = 'FFFFFF00'
                for cell in wbs['A']:
                    if cell.fill.start_color.index == highlight_color:
                        highlight_cell = cell
                        break
                    
                # Shift cells above the highlighted cell, NOT including the highlighted cell
                last_column = get_column_letter(wbs.max_column)
                wbs.move_range('A2:'+last_column+str(highlight_cell.row - 1), rows=(wbs.max_row-1))
                wbs.delete_rows(2, amount=highlight_cell.row - 2)
                
                
            # Delete the value in cell A1 since it is just for information
            wbs['A1'].value = None
            
            # Convert the caliper numbers in column A to degrees
            degree_scale = 360/(wbs.max_row - 1)
            i = 0
            for cell in wbs['A']:
                if cell.row == 1:
                    continue
                cell.value = degree_scale * i
                i += 1
                
            # Delete repeating columns based on the z value
            # Also delete columns where the z value reverts back
            # Note: Need to loop backwards to not affect the position of the column if a column is deleted
            z_start = wbs['B1'].value
            z_col_del = 0
            z_next = 0
            for i in reversed(range(2, wbs.max_column + 1)):
                # Convert from feet to inches
                i_col = get_column_letter(i) + '1'
                # i_col_next = get_column_letter(i + 1) + '1'
                wbs[i_col].value = (wbs[i_col].value - z_start) * 12
                # -------------------------------------------------------------< FIX HERE NEED TO ADD >= without errors happening!
                if wbs[i_col].value == z_next:
                # if wbs[i_col].value == wbs[i_col_next].value:
                    # Delete the column with the repeating value
                    # Delete the CURRENT column to not change the column order
                    z_col_del += 1
                    wbs.delete_cols(i,1)
                z_next = wbs[i_col].value
            
            # Convert the worksheet to a dataframe for faster numerical processing
            wbs_cols = np.array([[i.value for i in j] for j in wbs['B1:' + get_column_letter(wbs.max_column) + '1']]).flatten()
            wbs_indx = np.array([[i.value for i in j] for j in wbs['A2:A' + str(wbs.max_row)]]).flatten()
            wbs.delete_rows(1,1)
            wbs.delete_cols(1,1)
            wbs_data = wbs.values
            # Close the workbook
            wb.close()
            # Create DataFrame
            data = pd.DataFrame(data=wbs_data, columns=wbs_cols, index=wbs_indx)
            # ----- Continue using Pandas DataFrame -----
            theta_i   = data.index.to_numpy()
            radius_i  = data.to_numpy()
            z_i       = data.columns.to_numpy()
            rows_i,cols_i = radius_i.shape
        
        # ----- Load the Dent Registry -----
        # Only focus on specific columns of data
        reg_cols = ['Feature\nNo.','Event','Odometer [ft]','OD [in]','WT [in]','Depth in Inches [in]','Length [in]','Width [in]','o\'clock','Vendor Comment']
        # reg = pd.read_excel('template/dent_registry.xlsm', sheet_name=0, header=8)
        # DataFrame of Dent Registry for filtered columns
        dreg = reg.filter(reg_cols, axis=1)
        
        # The 'Feature\nNo.' is used for matching the ILI scan with the Dent Registry
        # The ILI scan data file can have multiple formats:
        # - 'Item 1234.xlsx'    Dent Number: Split text, digits, and extension
        # - '1234.csv'          Dent Number: Split digits from extension
        # - '1234_56.csv'       Odometer Reading: Split digits before and after _, and extension
        
        # Split from extension and remove leading/trailing spaces
        # This is ASSUMING there are NO other periods '.' in the string besides right before the file extension!
        
        feature_id = file.split('.')[0].strip()
        if feature_id.isnumeric() == True:
            feature_no = int(feature_id)
        elif not feature_id.lower().find('item') == -1:
            # Make lowercase, remove the leading text 'item', and remove leading/trailing spaces
            feature_no = int(feature_id.lower().replace('item','').strip())
        elif not feature_id.find('_') == -1:
            # Replace the '_' with a '.'
            feature_od = float(feature_id.replace('_','.'))
            # Find the corresponding feature_no from the Tally Sheet
            # The values in the Tally Sheet need to be rounded to 2 decimal places
            dreg_f = dreg.loc[round(dreg[reg_cols[2]],2) == feature_od] 
            feature_no = int(dreg_f[reg_cols[0]].to_numpy()[0])
            
        
        # The matching Dent Registry information on the ILI scan
        dreg_f = dreg.loc[dreg[reg_cols[0]] == feature_no]                          # Filtered dent registry for single feature
        reg_odometer  = float(dreg_f['Odometer [ft]'].to_numpy()[0])                # Odometer [ft]
        reg_pos_z     = (dreg_f['Odometer [ft]'].to_numpy()[0] - z_start)*12        # Convert to relative inches
        reg_pos_t1    = int(str(dreg_f['o\'clock'].to_numpy()[0]).split(':')[0])    # Orientation Hour
        reg_pos_t2    = int(str(dreg_f['o\'clock'].to_numpy()[0]).split(':')[1])    # Orientation Mintue
        reg_pos_t     = 30*reg_pos_t1 + 0.5*reg_pos_t2                              # Orientation in Deg
        reg_od        = dreg_f['OD [in]'].to_numpy()[0]                             # Outside diameter
        reg_wt        = dreg_f['WT [in]'].to_numpy()[0]                             # Wall thickness
        reg_ir        = reg_od/2 - reg_wt                                           # Internal radius
        reg_depth     = dreg_f['Depth in Inches [in]'].to_numpy()[0]                # Feature depth
        reg_length    = dreg_f['Length [in]'].to_numpy()[0]                         # Feature length
        reg_width     = dreg_f['Width [in]'].to_numpy()[0]                          # Feature width
        reg_event     = dreg_f['Event'].to_numpy()[0]                               # Feature event description
        reg_interaction = dreg_f['Vendor Comment'].to_numpy()[0]                    # Vendor Comments
        
        # circ_cal = caliper_spacing*rows_i
        # circ_rad = 2*np.pi*reg_ir
        # circ_err = 100*abs(circ_cal - circ_rad)/circ_rad
        # print('Pipe caliper spacing check by calculating circumference [in]:')
        # print('Using calipers = %.4f | Using radius = %.4f' % (circ_cal, circ_rad))
        # print('Circumference Error = %.2f %%' % (circ_err))
        
        # # ERROR CHECKS
        # if circ_err > 15:
        #     errorcheck = input("""ERROR: Caliper spacing difference exceeded 15%%. Please review the caliper spacing and the reported pipe internal radius before continuing this program. Type your selection and then press ENTER.\n
        #                        0 = ABORT | 1 = CONTINUE\n
        #                        User Input: """)
        #     if errorcheck == '0':
        #         sys.exit('PROGRAM ABORTED SUCCESSFULLY')
# =============================================================================
# DATA PROCESSING STEP: REMOVE BROKEN CALIPERS
# =============================================================================
    
        radius_cal = radius_i.copy()    
    
        if remove_calipers == True:
            
            total_time = time.time() - start_time
            print('%03d | ===== BROKEN CALIPERS ======' % (total_time))
            
            
            broken_cal = True
        
            while broken_cal == True:
                # Run through each caliper line, and see what the average, lower and upper bound of gradient compared to 
                # Axis = 1 because we only want the gradients along the Circumferential Direction (or columns direction)
                # spacing = np.arange(radius_i.shape[0])
                
                # Axis = 0 to take the gradient along the rows axis (one column, circumferential ring at a time)
                circ_grad = np.gradient(radius_cal, axis=0)
                # Determine the average per each Caliper Line (or row)
                # Axis = 1 to take the mean along the columns axis (one row, caliper at a time)
                circ_grad_avg = circ_grad.mean(axis=1)
                # Find the max average value and location
                # Since the positive slope will come AFTER the sudden drop, then the outlier value should be right before the max slope
                circ_grad_max_ind = np.argmax(circ_grad_avg) - 1
                
                if max(circ_grad_avg) > circ_grad_tol:
                    
                    # Demonstrate where it is at its worst
                    circ_grad_max_inds = np.unravel_index(np.argmin(radius_cal), radius_cal.shape)
                    # Before
                    plt.plot(radius_cal[:,circ_grad_max_inds[1]])
                    
                    # Replace the caliper with the average
                    radius_cal[circ_grad_max_ind,:] = (radius_cal[circ_grad_max_ind - 1,:] + radius_cal[circ_grad_max_ind + 1,:])/2
                    
                    # After
                    plt.plot(radius_cal[:,circ_grad_max_inds[1]])
                    
                    print('Replaced broken caliper at %.2f deg due to average gradient of %.4f' % (theta_i[circ_grad_max_ind], max(circ_grad_avg)))
                    
                else:
                    # Set to False to exit While loop
                    print('No broken caliper detected.')
                    broken_cal = False
                    
        
        radius_s = radius_cal.copy()
            
            
# =============================================================================
# DATA PROCESSING STEP 2: DATA SMOOTHING
# =============================================================================
        
        
        if data_smoothing == True:
            
            total_time = time.time() - start_time
            # print('%03d | ===== DATA SMOOTHING =======' % (total_time))
            
            # Check the axial spacing
            dz = 0
            for i in range(1,len(z_i)):
                dz += z_i[i] - z_i[i-1]
            dz = dz/i
            
            R0=reg_od/2     # nominal radius from ASME B31.8
            cwindow=5       # smoothing window (#points) for circumferential smoothing filter - must be an odd number (Default = 5)
            csmooth=0.0015  # circumferential smoothing parameter for splines (default 0.0015)
            axwindow=83     # smoothing window (#points) for axial smoothing filter
            asmooth=0.01    # axial smoothing parameter for splines
            
            #smoothed file output intervals (for ABAQUS, etc.)
            out_ax = 0.5    # axial length
            out_circ = 0.5  # circumferential length
            
            calwt = np.ones(theta_i.shape[0])
            axwt = np.ones(z_i.shape[0])
            
            axial = z_i.copy()
            rawdata = radius_s.copy()
            sensspac = np.deg2rad(theta_i)
            
            pipe_cpts = np.zeros(radius_cal.shape)
            pipe_apts = np.zeros(radius_cal.shape)
            
            #STEP #1 Circumferential Profiles Smoothing and Spline Functions
            for axpos, rad in enumerate(rawdata[0,:]):
                rad=rawdata[:,axpos] #isolate current 
                circ_filt = savgol_filter(rad, cwindow, 3, mode='wrap') #filter circumferential section
                circ_spl = splrep(sensspac, circ_filt, w=calwt, k=3, s=csmooth, per=1) #find tuples for current circumferential section
                circ_pts = splev(sensspac, circ_spl) #evalutate splines for current cross section
                pipe_cpts[:,axpos] = circ_pts #Variable to store all pipe points based on circumferential sections
                
        
            #STEP #2 Axial Profiles Smoothing and Spline Functions
            for cal, traces in enumerate(rawdata[:,0]):
                traces=rawdata[cal,:] #isolate current axial section
                ax_filt = savgol_filter(traces, axwindow, 3) #smooth current axial section
                ax_spl = splrep(axial, ax_filt, w=axwt, k=3, s=asmooth) #calcuate spline coefficients for current axial location
                ax_pts = splev(axial, ax_spl) #calculate axial points
                pipe_apts[cal,:] = ax_pts #Variable to store all pipe points based on axial cross sections      
        
        
            #STEP 3 - Create Weighted Average Profile from Axial and Circumfernetial Profiles
            err_circ = abs(pipe_cpts - rawdata) #calcualtes the variation for the circumferential profiles
            err_axial = abs(pipe_apts - rawdata) #calculates the variation for the axial profiles
            pipe_avgpts = (err_circ * pipe_cpts + err_axial * pipe_apts) / (err_circ + err_axial) #Average Weighted Positions based on magnitude of variations, greater the variation the higher the weight
        
            #---------------Create Soothed Output Data on Requested Interval from Pipe_Avgpts----------------
            #---------------Evaluate Strain on Desired Interval from Pipe Avgpts-----------------------------
        
            c_out_pts = math.ceil((math.pi*2.0*R0 / out_circ) / 4) * 4 #find circumferential interval closest to out_circ length on a multiple of four
            circ_out_int = np.linspace(0, 2*math.pi, c_out_pts, False) #create circumferential interval of radians
            pipe_out=np.empty((len(circ_out_int), len(axial)),) #create empty array to store variables
        
            a_out_pts = int(max(axial) / out_ax)
            ax_out_int = np.linspace(0, round(max(axial),0), a_out_pts, False) #create equally spaced axial points baed on a_ax_pts for smoothing
            f_pipe_out=np.empty((len(circ_out_int), len(ax_out_int)),) #create an empty array to store radii for smoothing
        
        
            #STEP 4: CALCULATE FINAL PROFILES - AXIAL FIRST
        
            for axpos, rad in enumerate(pipe_avgpts[0,:]):
                rad=pipe_avgpts[:,axpos] #isolate current cirumferential section
                out_spl = splrep(sensspac, rad, k=3, s=csmooth, per=1) #claculate interpolating spline
                out_pts = splev(circ_out_int, out_spl) #output circumferential sections on requested interval for smoothed output
                pipe_out[:,axpos] = out_pts #Variable to store all smoothed pipe points based on circumferential cross sections
        
            for cal, traces in enumerate(pipe_out[:,0]):
                traces=pipe_out[cal,:] #isolate current axial trace
                ax_filt = savgol_filter(traces, axwindow, 3) #smooth current axial section
                f_out_spl = splrep(axial, ax_filt, k=3, s=asmooth) #calcualte axial splines
                ax_pts = splev(ax_out_int, f_out_spl)
                f_pipe_out [cal,:] = ax_pts #final output points for smoothed file
                
            theta_f     = np.rad2deg(circ_out_int)
            z_f         = ax_out_int
            radius_f    = f_pipe_out
            
        else:
            # Do not perform data smoothing
            theta_f = theta_i.copy()
            radius_f = radius_i.copy()
            z_f = z_i.copy()
            
        # Output Data
        
        
        # Starting from the center of the dent, collect a perimeter
        # Output data must be in (100,100) dimensions
        long_len = 100
        circ_len = 100
        # Center at (101,81)
        # radius_o = np.zeros([long_len*2 + 1,circ_len*2 + 1])
        # Location of the lowest point
        min_ind = np.unravel_index(np.argmin(radius_f), radius_f.shape)
        print('OT-%03d | min_ind = [%d,%d]' % (time.time() - overall_time, min_ind[0], min_ind[1]))
        # Copy and paste from the center to the limits
        RLB = min_ind[0] - int(long_len/2)
        CLB = min_ind[1] - int(circ_len/2)
        if RLB < 0: RLB = 0
        if CLB < 0: CLB = 0
        radius_o = radius_f[RLB:min_ind[0] + int(long_len/2),
                            CLB:min_ind[1] + int(circ_len/2)]
        # min_ind_o = np.unravel_index(np.argmin(radius_o), radius_o.shape)
        # Normalize the Radial values
        radius_o = (radius_o - reg_ir)/reg_ir
        # If radius_f was below the desired dimensions, then add empty columns
        if radius_o.shape[1] < long_len:
            while radius_o.shape[1] < long_len:
                # Add column on both sides at the same time
                radius_o = np.hstack((radius_o, np.zeros([radius_o.shape[0],1])))
                radius_o = np.hstack((np.zeros([radius_o.shape[0],1]), radius_o))
            # Check if the dimensions match now
            if radius_o.shape[1] > long_len:
                radius_o = np.delete(radius_o, 0, 1)
        if radius_o.shape[0] < circ_len:
            while radius_o.shape[0] < circ_len:
                # Add row on both sides at the same time
                radius_o = np.vstack((radius_o, np.zeros([1,radius_o.shape[1]])))
                radius_o = np.vstack((np.zeros([1,radius_o.shape[1]]), radius_o))
            # Check if the dimensions match now
            if radius_o.shape[0] > circ_len:
                radius_o = np.delete(radius_o, 0, 0)
                
        plt.figure()
        plt.xticks([])
        plt.yticks([])
        plt.imshow(radius_o, cmap=plt.cm.binary)
        plt.xlabel(str(feature_no))
        plt.colorbar()
        plt.grid(False)
        plt.show()
        
        output_file = 'results/' + str(feature_no) + '.csv'
        np.savetxt(output_file, radius_o, delimiter=",")
        
        radius_o = np.expand_dims(radius_o, axis=0)
        if first_file == True:
            # do nothing
            first_file = False
            radius_out = radius_o.copy()
            scf_out = scf_df['SCF'].loc[scf_df['Dent #'] == feature_no].to_numpy()
        else:
            radius_out = np.concatenate((radius_out,radius_o),axis=0)
            scf_out = np.concatenate((scf_out, scf_df['SCF'].loc[scf_df['Dent #'] == feature_no].to_numpy()))
        
        # =============================================================================
        # STEP 8: TIME
        # =============================================================================
    
        # Export the dent profile contour values
        total_time = time.time() - start_time
        time_min, time_sec = divmod(total_time, 60)
        
        print('OT-%03d | =========== END ============' % (time.time() - overall_time))
    
# Overall Finished
# Save the resultant radius_out and scf_out data
np.save('radius_out',radius_out)
np.save('scf_out', scf_out)

o_total_time = time.time() - overall_time
o_time_m, o_time_s = divmod(o_total_time, 60)
o_time_h, o_time_m = divmod(o_time_m, 60)
print('Total overall time of %.0f:%.0f:%.0f' % (o_time_h, o_time_m, o_time_s))
