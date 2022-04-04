# -*- coding: utf-8 -*-
"""
Regression ML Predictions
Author: Emmanuel Valencia
Created on: 02/10/2022
Updated on: 04/04/2022
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
# All scikit-learn related modules
from sklearn.model_selection import train_test_split
import sklearn.metrics as skm
# All TensorFlow related modules
import tensorflow as tf
# import tensorflow_addons as tfa
from tensorflow.keras import layers, models
# import pathlib
from openpyxl import load_workbook
import datetime
import random
import time
import os
import sys

# Will need to have Input Data as 3-dimensional ILI layers
# - var1: number of ILI data files
# - var2: dimension in the longitudinal (along pipe axis)
# - var3: dimension in the lateral (along pipe circumference)
# To simplify the preprocessing, will use square resolution, meaning that the
# unit of length in var2 should be equal to that of var3. This way the length
# is not warped in any direction (more resolution in one direction). However,
# this does not mean that values var2 and var3 need to be equal. The resultant
# image should be like a rectangle.

# =============================================================================
# ADJUSTED PARAMETERS
# =============================================================================

test_iteration  = 34
model_number    = test_iteration
data_used       = 'March 2nd'

# Hyperparameters
conv2D_shapes       = [64,64,64] # [32,64,128,256]
conv2D_windows      = [(3,3),(3,3),(3,3)] # [(3,3),(3,3),(3,3),(3,3)]
conv2D_activations  = ['relu','relu','relu'] # ['relu','relu','relu','relu']
dense_shapes        = [120,60,1]
dense_activations   = ['relu','relu']
compile_optimizer   = 'adam'
compile_loss        = 'mse'
compile_metrics     = ['mae']
fit_epochs          = 50

# =============================================================================
# FIXED PARAMETERS
# =============================================================================

time_start = time.time()

# Input Data dimensions
axial_len   = 200
circ_len    = 200
shape       = [axial_len, circ_len]

metrics_path        = 'metrics/'
metrics_name        = 'metrics_regression.xlsx'
training_data_path  = 'training_data'
results_path        = 'results/'
models_path         = 'models/'

# =============================================================================
# FUNCTIONS
# =============================================================================

def load_training_data(training_data_path):
    # Load Training Data
    image_data_path = []
    label_data_path = []

    for subdir, dirs, files in os.walk(training_data_path):
        for file in files:
            if 'radius' in file:
                image_data_path.append(os.path.join(subdir, file))
            if 'SCF' in file:
                label_data_path.append(os.path.join(subdir, file))

    # Load ALL of the data .npy files
    if len(image_data_path) != len(label_data_path):
        sys.exit()

    for i in range(len(image_data_path)):
        if i == 0:
            image_data = np.load(image_data_path[i])
            label_data = np.load(label_data_path[i])
            continue

        image_data = np.concatenate((image_data, np.load(image_data_path[i])), axis=0)
        label_data = np.concatenate((label_data, np.load(label_data_path[i])), axis=0)

    return image_data, label_data

def regressor_model(shape):
    model = models.Sequential()
    # Layer 1
    model.add(layers.Conv2D(conv2D_shapes[0], conv2D_windows[0], activation=conv2D_activations[0], input_shape=(shape[0], shape[1], 1)))
    model.add(layers.MaxPooling2D((2,2)))
    # Layer 2
    model.add(layers.Conv2D(conv2D_shapes[1], conv2D_windows[1], activation=conv2D_activations[1]))
    model.add(layers.MaxPooling2D((2,2)))
    # Layer 3
    model.add(layers.Conv2D(conv2D_shapes[2], conv2D_windows[2], activation=conv2D_activations[2]))
    model.add(layers.MaxPooling2D((2,2)))
    # Layer 4
    # model.add(layers.Conv2D(conv2D_shapes[3], conv2D_windows[3], activation=conv2D_activations[3]))
    # model.add(layers.MaxPooling2D((2,2)))

    # Add dense layers on top
    model.add(layers.Flatten())
    model.add(layers.Dense(dense_shapes[0], activation=dense_activations[0]))
    model.add(layers.Dense(dense_shapes[1], activation=dense_activations[1]))
    model.add(layers.Dense(dense_shapes[2]))

    # Compile and train model
    model.compile(optimizer=compile_optimizer, loss=compile_loss, metrics=compile_metrics)

    return model

def plot_image_prediction(j, image, label_true, label_predicted):
    plt.grid(False)
    plt.xticks([])
    plt.yticks([])
    plt.imshow(image, cmap=plt.cm.RdYlGn)
    # Negative Error Percentage means that the model UNDER predicted the SCF value
    error_perc = (label_predicted - label_true)/label_true * 100
    plt.xlabel("#%d | %.3f (%.3f) | %.0f%%" % (j, label_predicted, label_true, error_perc))

def metrics(y_true, y_pred):
    # R^2 / Adjusted R^2
    # Measures how much variability in the dependent variable can be explained by the model.
    # It is the square of the Correlation Coefficient, R.
    # Relative measure of how well the model fits dependent variables.
    r2 = skm.r2_score(y_true, y_pred)

    # Mean Square Error (MSE) / Root Mean Square Error (RMSE)
    # Absolute measure of the goodness for the fit. Sum of square of error.
    # Gives larger penalization to big prediction error by squaring it.
    mse = skm.mean_squared_error(y_true, y_pred)
    rmse = np.sqrt(mse)

    # Mean Absolute Error (MAE)
    # Sum of the absolute value of error. More direct representation of sum of error terms.
    # Treats all errors the same.
    mae = skm.mean_absolute_error(y_true, y_pred)

    # Max Error
    max_err = skm.max_error(y_true, y_pred)

    return r2, mse, rmse, mae, max_err

def metrics_data_print():
    # Metrics Values
    # Metrics Values in list organization
    metrics_data = [str(datetime.datetime.now()),   # Time stamp
                    wbMs_row - 1,                   # Run number
                    data_used,
                    str(model_number),
                    str(conv2D_shapes),
                    str(conv2D_windows),
                    str(conv2D_activations),
                    str(dense_shapes),
                    str(dense_activations),
                    compile_optimizer,
                    compile_loss,
                    compile_metrics[0],
                    label_data.shape[0],
                    label_train.shape[0],
                    train_loss[-1],
                    train_metric[-1],
                    training_time_stop - training_time_start,   # Training time
                    fit_epochs,
                    label_val.shape[0],
                    val_loss,
                    val_metric,
                    r2,
                    mse,
                    rmse,
                    mae,
                    max_error,
                    max_error_perc,
                    time.time() - time_iteration]       # Total time
    return metrics_data

# =============================================================================
# SCRIPT ITERATIONS
# =============================================================================

# run_inf = 0
# while run_inf < 100:
#     run_inf = run_inf + 1

# Current Time
test_iteration  = 1
time_iteration  = time.time()
time_current    = time.time() - time_start
time_m, time_s  = divmod(time_current, 60)
time_h, time_m  = divmod(time_m, 60)
print('========== START ==========')
print('%.0f:%.0f:%.0f | Program began execution.' % (time_h, time_m, time_s))

# Load All Data
image_data, label_data = load_training_data(training_data_path)

# Split the data into Training, Validation, and Test sets using 70:20:10 split
image_trainval, image_test, label_trainval, label_test = train_test_split(image_data, label_data, test_size=0.10, random_state=42) # Temporarily use a random state for reproducibility
image_train, image_val, label_train, label_val = train_test_split(image_trainval, label_trainval, test_size=0.18) # 0.20 * 0.90 = 0.18

# Observe the distribution of data
# SCF Distribution for Training Data
plt.figure(figsize=(16,8))
plt.subplot(1,3,1)
plt.hist(np.ravel(label_train), bins=20, range=[1,11])
plt.xlabel('Training Data SCF')
plt.ylabel('Frequency')
# SCF Distribution for Validation Data
plt.subplot(1,3,2)
plt.hist(np.ravel(label_val), bins=20, range=[1,11])
plt.xlabel('Validation Data SCF')
plt.ylabel('Frequency')
# SCF Distribution for Validation Data
plt.subplot(1,3,3)
plt.hist(np.ravel(label_test), bins=20, range=[1,11])
plt.xlabel('Testing Data SCF')
plt.ylabel('Frequency')
plt.savefig(results_path + str(test_iteration) + '_train_val_test_dist.png')

# Test 9 random images: Generate a random set of test images and labels
rand_list_test = random.sample(range(image_train.shape[0]), 9)
plt.figure(figsize=(10,10))
for i, j in enumerate(rand_list_test):
    plt.subplot(3,3,i+1)
    plt.xticks([])
    plt.yticks([])
    plt.grid(False)
    plt.imshow(image_train[j], cmap=plt.cm.RdYlGn)
    plt.xlabel(round(label_train[j],3))
    plt.colorbar()
plt.show()

model = regressor_model(shape)  # Create the regression ML model
model.summary()                 # Print the current model information

training_time_start = time.time()
history = model.fit(image_train, label_train, epochs=fit_epochs,
                    validation_data=(image_val, label_val))
training_time_stop = time.time()

# Model Training Evaluation
train_loss = history.history['loss']
train_metric = history.history[compile_metrics[0]]

# Model Validation Evaluation
val_loss, val_metric = model.evaluate(image_val, label_val, verbose=2)
print('%.0f:%.0f:%.0f | Run %03d | Train Val Loss = %.3f | Train Val Metric = %.3f' % (time_h, time_m, time_s, test_iteration, val_loss, val_metric))

# Evaluate the model: Metrics
plt.figure(figsize=(8,6))
plt.subplot(2,1,1)
plt.plot(history.history[compile_metrics[0]], label='Training Metric')
plt.plot(history.history['val_' + compile_metrics[0]], label='Validation Metric')
plt.xlabel('Epoch')
plt.ylabel('Metrics = ' + compile_metrics[0])
plt.ylim([0, 1])
plt.legend(loc='upper right')

# Evaluate the model: Losses
plt.subplot(2,1,2)
plt.plot(history.history['loss'], label='Training Loss')
plt.plot(history.history['val_loss'], label='Validation Loss')
plt.xlabel('Epoch')
plt.ylabel('Loss = ' + compile_loss)
plt.ylim([0, 1])
plt.legend(loc='upper right')
plt.savefig(results_path + str(test_iteration) + '_train_val.png')

# =============================================================================
# VALIDATION
# =============================================================================

# Make predictions
predictions_val = model.predict(image_val)
residuals_val = np.ravel(predictions_val) - label_val

# Plot some predictions
num_rows = 3
num_cols = 3
num_images = num_rows * num_cols
plt.figure(figsize=(8,8))
# Generate a random set of test images and labels
rand_list = random.sample(range(predictions_val.shape[0]), num_images)
# rand_list = [267, 695, 617, 214, 545, 353, 165, 463, 27]
for i, j in enumerate(rand_list):
    plt.subplot(num_rows, num_cols, i + 1)
    plot_image_prediction(j, image_val[j], label_val[j], predictions_val[j][0])
plt.tight_layout()
plt.show()
# plt.savefig(results_path + str(test_iteration) + '_predictions.png')

# SCF Values Unity
unity_line = np.linspace(0,15,100)
plt.figure(figsize=(16,8))
plt.subplot(1,2,1)
plt.plot(unity_line, unity_line, '-k')
plt.plot(label_val, np.ravel(predictions_val), 'or')
plt.xlabel('Truth Data (SCF)')
plt.ylabel('Predicted Data (SCF)')
plt.xlim([0,14])
plt.ylim([0,14])
# Unity Plot Distribution
plt.subplot(1,2,2)
plt.hist(np.ravel(predictions_val), bins=20, range=[1,11])
plt.xlabel('Predicted Data (SCF)')
plt.ylabel('Frequency')
plt.savefig(results_path + str(test_iteration) + '_scf_values.png')

# Residual Error Unity
plt.figure(figsize=(16,8))
plt.subplot(1,2,1)
plt.plot(unity_line, np.linspace(0,0,100), '-k')
plt.plot(label_val, residuals_val, 'or')
plt.xlabel('Truth Data (SCF)')
plt.ylabel('Residual Error (SCF)')
plt.xlim([0,14])
plt.ylim([-8,8])
# Residual Error Distribution
plt.subplot(1,2,2)
plt.hist(residuals_val, bins=20, range=[-4,4])
plt.xlabel('Residual Error (SCF)')
plt.ylabel('Frequency')
plt.savefig(results_path + str(test_iteration) + '_scf_residuals.png')

# =============================================================================
# METRICS
# =============================================================================

# Save the Metrics to the metrics_regression.xlsx Workbook

# Metrics
r2, mse, rmse, mae, max_err = metrics(label_val, predictions_val)
max_error       = skm.max_error(label_val, predictions_val)
max_error_perc  = max((np.ravel(predictions_val) - label_val)/label_val)

metrics_file    = metrics_path + metrics_name
wbM             = load_workbook(metrics_file)
wbM_sn          = wbM.sheetnames
wbMs            = wbM[wbM_sn[0]]
wbMs_row        = wbMs.max_row

metrics_labels  = wbMs[2]
metrics_labels  = [cell.value for cell in metrics_labels]

metrics_data    = metrics_data_print()

# Write the metric values to the Excel workbook
for i, value in enumerate(metrics_data):
    wbMs.cell(row=wbMs_row+1, column=i+1, value=value)
wbM.save(metrics_file)
wbM.close()     # Close the Results Workbook

# Save the current model
model.save(models_path + str(test_iteration) + '_model')

# Total time
time_current    = time.time() - time_start
time_m, time_s  = divmod(time_current, 60)
time_h, time_m  = divmod(time_m, 60)
print('%.0f:%.0f:%.0f | Overall total time.' % (time_h, time_m, time_s))
print('=========== END ===========')
